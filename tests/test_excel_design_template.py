from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from datamapx.excel_design_template import SHEET_ORDER, write_template

EXAMPLE_DIR = Path(__file__).resolve().parents[1] / "examples" / "08_excel_design"
SHEETS_DIR = EXAMPLE_DIR / "sheets"
TEMPLATE_PATH = EXAMPLE_DIR / "datamapx_design_template.xlsx"


def _read_csv_headers(path: Path) -> list[str]:
    with path.open("r", encoding="utf-8", newline="") as file:
        reader = csv.reader(file)
        return next(reader)


def _read_workbook_headers(workbook, sheet_name: str) -> list[str]:
    worksheet = workbook[sheet_name]
    return [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]


def test_write_template_matches_sheet_csvs(tmp_path: Path) -> None:
    output_path = tmp_path / "datamapx_design_template.xlsx"
    write_template(output_path, SHEETS_DIR)

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == SHEET_ORDER
    for sheet_name in SHEET_ORDER:
        assert _read_workbook_headers(workbook, sheet_name) == _read_csv_headers(
            SHEETS_DIR / f"{sheet_name}.csv"
        )


def test_committed_template_matches_sheet_csvs() -> None:
    workbook = load_workbook(TEMPLATE_PATH)
    assert workbook.sheetnames == SHEET_ORDER
    for sheet_name in SHEET_ORDER:
        assert _read_workbook_headers(workbook, sheet_name) == _read_csv_headers(
            SHEETS_DIR / f"{sheet_name}.csv"
        )
