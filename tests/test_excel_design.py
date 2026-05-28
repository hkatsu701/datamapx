from __future__ import annotations

import csv
import json
import shutil
from pathlib import Path

from openpyxl import load_workbook

from datamapx.excel_design import (
    format_design_result,
    validate_design_workbook,
    write_design_errors_csv,
    write_design_summary_json,
)

EXAMPLE_DIR = Path(__file__).resolve().parents[1] / "examples" / "08_excel_design"
TEMPLATE_PATH = EXAMPLE_DIR / "datamapx_design_template.xlsx"


def _copy_template(tmp_path: Path, filename: str = "design.xlsx") -> Path:
    target = tmp_path / filename
    shutil.copy2(TEMPLATE_PATH, target)
    return target


def _load_workbook(path: Path):
    return load_workbook(path)


def test_validate_design_template_success() -> None:
    result = validate_design_workbook(TEMPLATE_PATH)

    assert result.valid
    assert result.project_name == "invoice_migration"
    assert result.sheet_count == 16
    assert result.job_count == 2
    assert result.enabled_job_count == 2
    assert "Design is valid:" in format_design_result(result)


def test_validate_design_template_writes_summary_and_empty_errors(tmp_path: Path) -> None:
    result = validate_design_workbook(TEMPLATE_PATH)
    summary_path = tmp_path / "design-summary.json"
    errors_path = tmp_path / "design-errors.csv"

    write_design_summary_json(summary_path, result)
    write_design_errors_csv(errors_path, result.errors)

    summary = json.loads(summary_path.read_text(encoding="utf-8"))
    assert summary["project_name"] == "invoice_migration"
    assert summary["sheet_count"] == 16
    assert summary["job_count"] == 2
    assert summary["enabled_job_count"] == 2
    assert summary["valid"] is True
    with errors_path.open("r", encoding="utf-8", newline="") as file:
        rows = list(csv.reader(file))
    assert rows == [["sheet", "row_number", "column", "message"]]


def test_missing_required_sheet_fails(tmp_path: Path) -> None:
    workbook_path = _copy_template(tmp_path)
    workbook = _load_workbook(workbook_path)
    workbook.remove(workbook["runtime"])
    workbook.save(workbook_path)

    result = validate_design_workbook(workbook_path)

    assert not result.valid
    assert any(error.sheet == "runtime" and "missing" in error.message for error in result.errors)


def test_missing_required_column_fails(tmp_path: Path) -> None:
    workbook_path = _copy_template(tmp_path)
    workbook = _load_workbook(workbook_path)
    workbook["jobs"].delete_cols(1)
    workbook.save(workbook_path)

    result = validate_design_workbook(workbook_path)

    assert not result.valid
    assert any(
        error.sheet == "jobs" and error.row_number == 1 and error.column == "job_id"
        for error in result.errors
    )


def test_jobs_optional_columns_may_be_omitted(tmp_path: Path) -> None:
    workbook_path = _copy_template(tmp_path)
    workbook = _load_workbook(workbook_path)
    workbook["jobs"].delete_cols(6, 2)
    workbook.save(workbook_path)

    result = validate_design_workbook(workbook_path)

    assert result.valid
    assert result.job_count == 2


def test_duplicate_job_id_fails(tmp_path: Path) -> None:
    workbook_path = _copy_template(tmp_path)
    workbook = _load_workbook(workbook_path)
    worksheet = workbook["jobs"]
    worksheet.append(
        [
            "01_merge_customers",
            "merge",
            3,
            True,
            "merge_customers_copy",
            "",
            "Duplicate job",
        ]
    )
    workbook.save(workbook_path)

    result = validate_design_workbook(workbook_path)

    assert not result.valid
    assert any("duplicate job_id" in error.message for error in result.errors)


def test_unsupported_job_type_fails(tmp_path: Path) -> None:
    workbook_path = _copy_template(tmp_path)
    workbook = _load_workbook(workbook_path)
    workbook["jobs"]["B2"] = "export"
    workbook.save(workbook_path)

    result = validate_design_workbook(workbook_path)

    assert not result.valid
    assert any(
        error.sheet == "jobs" and error.row_number == 2 and error.column == "job_type"
        for error in result.errors
    )


def test_invalid_order_fails(tmp_path: Path) -> None:
    workbook_path = _copy_template(tmp_path)
    workbook = _load_workbook(workbook_path)
    workbook["jobs"]["C2"] = 0
    workbook.save(workbook_path)

    result = validate_design_workbook(workbook_path)

    assert not result.valid
    assert any(
        error.sheet == "jobs" and error.row_number == 2 and error.column == "order"
        for error in result.errors
    )


def test_unknown_depends_on_fails(tmp_path: Path) -> None:
    workbook_path = _copy_template(tmp_path)
    workbook = _load_workbook(workbook_path)
    workbook["jobs"]["F3"] = "missing_job"
    workbook.save(workbook_path)

    result = validate_design_workbook(workbook_path)

    assert not result.valid
    assert any("unknown job 'missing_job'" in error.message for error in result.errors)


def test_dependency_cycle_fails(tmp_path: Path) -> None:
    workbook_path = _copy_template(tmp_path)
    workbook = _load_workbook(workbook_path)
    workbook["jobs"]["F2"] = "02_migrate_invoices"
    workbook["jobs"]["F3"] = "01_merge_customers"
    workbook.save(workbook_path)

    result = validate_design_workbook(workbook_path)

    assert not result.valid
    assert any("dependency cycle detected" in error.message for error in result.errors)


def test_unknown_detail_job_id_fails(tmp_path: Path) -> None:
    workbook_path = _copy_template(tmp_path)
    workbook = _load_workbook(workbook_path)
    worksheet = workbook["merge_inputs"]
    worksheet.append(["unknown_job", "base", "x", "./input/x.csv", "utf-8-sig", ",", True, "id"])
    workbook.save(workbook_path)

    result = validate_design_workbook(workbook_path)

    assert not result.valid
    assert any(
        error.sheet == "merge_inputs"
        and error.row_number == 4
        and error.column == "job_id"
        and "unknown job_id" in error.message
        for error in result.errors
    )


def test_enabled_merge_job_requires_merge_rows(tmp_path: Path) -> None:
    workbook_path = _copy_template(tmp_path)
    workbook = _load_workbook(workbook_path)
    workbook["merge_inputs"].delete_rows(2, amount=2)
    workbook.save(workbook_path)

    result = validate_design_workbook(workbook_path)

    assert not result.valid
    assert any(
        error.sheet == "jobs"
        and error.row_number == 2
        and error.column == "job_id"
        and "merge_inputs" in error.message
        for error in result.errors
    )


def test_enabled_migration_job_requires_migration_rows(tmp_path: Path) -> None:
    workbook_path = _copy_template(tmp_path)
    workbook = _load_workbook(workbook_path)
    workbook["outputs"].delete_rows(2)
    workbook.save(workbook_path)

    result = validate_design_workbook(workbook_path)

    assert not result.valid
    assert any(
        error.sheet == "jobs"
        and error.row_number == 3
        and error.column == "job_id"
        and "outputs" in error.message
        for error in result.errors
    )
