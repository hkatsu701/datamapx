"""Validation for the standard DataMapX Excel design workbook."""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


class DesignWriteError(Exception):
    """Raised when design validation artifacts cannot be written."""


@dataclass(frozen=True)
class DesignErrorRow:
    """A single validation error in the Excel design workbook."""

    sheet: str
    row_number: int | None
    column: str | None
    message: str


@dataclass(frozen=True)
class DesignJob:
    """A validated job row."""

    job_id: str
    job_type: str
    order: int
    enabled: bool
    config_name: str
    depends_on: list[str]
    row_number: int


@dataclass(frozen=True)
class DesignValidationResult:
    """Validation result for one Excel design workbook."""

    design_path: str
    project_name: str | None
    sheet_count: int
    job_count: int
    enabled_job_count: int
    sheet_names: list[str]
    jobs: dict[str, DesignJob]
    errors: list[DesignErrorRow]

    @property
    def valid(self) -> bool:
        return not self.errors

    def summary_payload(self) -> dict[str, Any]:
        return {
            "design_path": self.design_path,
            "project_name": self.project_name,
            "sheet_count": self.sheet_count,
            "job_count": self.job_count,
            "enabled_job_count": self.enabled_job_count,
            "valid": self.valid,
            "error_count": len(self.errors),
            "sheet_names": self.sheet_names,
        }


REQUIRED_SHEETS = [
    "project",
    "jobs",
    "merge_inputs",
    "merge_rules",
    "migration_inputs",
    "input_schema",
    "references",
    "reference_schema",
    "derived",
    "outputs",
    "mappings",
    "validations",
    "filters",
    "checks",
    "error_handling",
    "runtime",
]

REQUIRED_COLUMNS: dict[str, list[str]] = {
    "project": ["project_name", "description", "default_encoding", "default_delimiter"],
    "jobs": ["job_id", "job_type", "order", "enabled", "config_name"],
    "merge_inputs": [
        "job_id",
        "role",
        "input_name",
        "path",
        "encoding",
        "delimiter",
        "header",
        "key",
    ],
    "merge_rules": [
        "job_id",
        "output_column",
        "rule_type",
        "source_input",
        "source_column",
        "fallback_input",
        "fallback_column",
        "description",
    ],
    "migration_inputs": ["job_id", "input_name", "path", "encoding", "delimiter", "header"],
    "input_schema": [
        "job_id",
        "input_name",
        "field",
        "source_columns",
        "type",
        "required",
        "normalize",
        "true_values",
        "false_values",
    ],
    "references": [
        "job_id",
        "reference_name",
        "path",
        "encoding",
        "delimiter",
        "header",
        "key",
        "on_duplicate",
    ],
    "reference_schema": [
        "job_id",
        "reference_name",
        "field",
        "source_columns",
        "type",
        "required",
        "normalize",
    ],
    "derived": [
        "job_id",
        "field",
        "rule_type",
        "expression",
        "source",
        "values",
        "map_rules",
        "when_rules",
        "lookup_reference",
        "lookup_key",
        "lookup_value",
        "lookup_default",
    ],
    "outputs": [
        "job_id",
        "output_name",
        "path",
        "encoding",
        "delimiter",
        "header",
        "newline",
        "if_exists",
        "columns",
    ],
    "mappings": [
        "job_id",
        "output_name",
        "output_column",
        "rule_type",
        "source",
        "value",
        "lookup_reference",
        "lookup_key",
        "lookup_value",
        "lookup_default",
        "expression",
        "when_rules",
    ],
    "validations": ["job_id", "scope", "field", "rule", "value", "values"],
    "filters": ["job_id", "filter_type", "condition", "reason"],
    "checks": ["job_id", "name", "rule"],
    "error_handling": [
        "job_id",
        "on_validation_error",
        "on_lookup_missing",
        "on_transform_error",
        "max_errors",
        "error_output",
        "skipped_output",
        "include_original_row",
    ],
    "runtime": ["job_id", "run_id", "log_dir", "log_level", "summary_output"],
}

OPTIONAL_COLUMNS: dict[str, list[str]] = {
    "jobs": ["depends_on", "description"],
}

MERGE_REQUIRED_DETAIL_SHEETS = ["merge_inputs", "merge_rules"]
MIGRATION_REQUIRED_DETAIL_SHEETS = [
    "migration_inputs",
    "input_schema",
    "outputs",
    "mappings",
    "error_handling",
    "runtime",
]


def validate_design_workbook(design_path: str | Path) -> DesignValidationResult:
    """Validate a standard Excel design workbook."""

    path = Path(design_path)
    errors: list[DesignErrorRow] = []
    jobs: dict[str, DesignJob] = {}
    sheet_rows: dict[str, list[tuple[int, dict[str, Any]]]] = {}

    workbook = _load_workbook(path, errors)
    if workbook is None:
        return DesignValidationResult(
            design_path=str(path),
            project_name=None,
            sheet_count=0,
            job_count=0,
            enabled_job_count=0,
            sheet_names=[],
            jobs={},
            errors=errors,
        )

    sheet_names = list(workbook.sheetnames)
    sheet_count = len(sheet_names)
    for sheet_name in REQUIRED_SHEETS:
        if sheet_name not in workbook.sheetnames:
            errors.append(
                DesignErrorRow(
                    sheet=sheet_name,
                    row_number=None,
                    column=None,
                    message="required sheet is missing",
                )
            )
            continue

        worksheet = workbook[sheet_name]
        headers = _read_headers(worksheet)
        expected_headers = REQUIRED_COLUMNS[sheet_name]
        optional_headers = OPTIONAL_COLUMNS.get(sheet_name, [])
        allowed_headers = expected_headers + optional_headers
        missing_headers = [column for column in expected_headers if column not in headers]
        extra_headers = [column for column in headers if column not in allowed_headers]
        if missing_headers or extra_headers:
            for column in missing_headers:
                errors.append(
                    DesignErrorRow(
                        sheet=sheet_name,
                        row_number=1,
                        column=column,
                        message="required column is missing",
                    )
                )
            for column in extra_headers:
                errors.append(
                    DesignErrorRow(
                        sheet=sheet_name,
                        row_number=1,
                        column=column,
                        message="unexpected column",
                    )
                )

        rows = _read_sheet_rows(worksheet, headers)
        sheet_rows[sheet_name] = rows

    for sheet_name in workbook.sheetnames:
        if sheet_name not in REQUIRED_SHEETS:
            errors.append(
                DesignErrorRow(
                    sheet=sheet_name,
                    row_number=None,
                    column=None,
                    message="unexpected sheet",
                )
            )

    project_name = _extract_project_name(sheet_rows.get("project", []), errors)
    jobs = _validate_jobs(sheet_rows.get("jobs", []), errors)
    _validate_job_dependencies(jobs, errors)
    _validate_detail_rows(sheet_rows, jobs, errors)
    _validate_required_detail_rows(sheet_rows, jobs, errors)

    return DesignValidationResult(
        design_path=str(path),
        project_name=project_name,
        sheet_count=sheet_count,
        job_count=len(jobs),
        enabled_job_count=sum(1 for job in jobs.values() if job.enabled),
        sheet_names=sheet_names,
        jobs=jobs,
        errors=errors,
    )


def write_design_summary_json(path: Path, result: DesignValidationResult) -> Path:
    """Write the design summary JSON payload."""

    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(
            json.dumps(result.summary_payload(), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except OSError as exc:
        raise DesignWriteError(f"{path}: cannot write design summary: {exc}") from exc
    return path


def write_design_errors_csv(path: Path, errors: list[DesignErrorRow]) -> Path:
    """Write structured design errors to CSV."""

    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8", newline="") as file:
            writer = csv.DictWriter(file, fieldnames=["sheet", "row_number", "column", "message"])
            writer.writeheader()
            for error in errors:
                writer.writerow(
                    {
                        "sheet": error.sheet,
                        "row_number": error.row_number if error.row_number is not None else "",
                        "column": error.column or "",
                        "message": error.message,
                    }
                )
    except OSError as exc:
        raise DesignWriteError(f"{path}: cannot write design errors: {exc}") from exc
    return path


def format_design_result(result: DesignValidationResult) -> str:
    """Return a human-readable validation summary."""

    if result.valid:
        lines = [
            f"Design is valid: {result.design_path}",
            f"Project: {result.project_name}",
            f"Sheets: {result.sheet_count}",
            f"Jobs: {result.job_count}",
            f"Enabled jobs: {result.enabled_job_count}",
        ]
        return "\n".join(lines)

    lines = [f"Design is invalid: {result.design_path}"]
    for error in result.errors:
        location = error.sheet
        if error.row_number is not None:
            location += f" row {error.row_number}"
        if error.column is not None:
            location += f" column {error.column}"
        lines.append(f"- {location}: {error.message}")
    return "\n".join(lines)


def _load_workbook(path: Path, errors: list[DesignErrorRow]):
    if path.suffix.lower() != ".xlsx":
        errors.append(
            DesignErrorRow(
                sheet="workbook",
                row_number=None,
                column=None,
                message="design workbook must be a .xlsx file",
            )
        )
        return None

    try:
        return load_workbook(path, data_only=True)
    except FileNotFoundError:
        errors.append(
            DesignErrorRow(
                sheet="workbook",
                row_number=None,
                column=None,
                message=f"{path}: workbook not found",
            )
        )
    except InvalidFileException as exc:
        errors.append(
            DesignErrorRow(
                sheet="workbook",
                row_number=None,
                column=None,
                message=str(exc),
            )
        )
    except OSError as exc:
        errors.append(
            DesignErrorRow(
                sheet="workbook",
                row_number=None,
                column=None,
                message=f"{path}: cannot read workbook: {exc}",
            )
        )
    return None


def _read_headers(worksheet) -> list[str]:
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [str(value).strip() if isinstance(value, str) else value for value in header_row]
    return _trim_trailing_empty(headers)


def _trim_trailing_empty(values: list[Any]) -> list[Any]:
    trimmed = list(values)
    while trimmed and trimmed[-1] in (None, ""):
        trimmed.pop()
    return trimmed


def _read_sheet_rows(worksheet, headers: list[str]) -> list[tuple[int, dict[str, Any]]]:
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        row = {}
        row_has_value = False
        for index, header in enumerate(headers):
            value = values[index] if index < len(values) else None
            row[header] = value
            if value not in (None, ""):
                row_has_value = True
        if row_has_value:
            rows.append((row_number, row))
    return rows


def _extract_project_name(
    rows: list[tuple[int, dict[str, Any]]],
    errors: list[DesignErrorRow],
) -> str | None:
    if not rows:
        errors.append(
            DesignErrorRow(
                sheet="project",
                row_number=None,
                column="project_name",
                message="project sheet requires at least one data row",
            )
        )
        return None

    _, row = rows[0]
    value = _clean_text(row.get("project_name"))
    if not value:
        errors.append(
            DesignErrorRow(
                sheet="project",
                row_number=rows[0][0],
                column="project_name",
                message="project_name is required",
            )
        )
        return None
    return value


def _validate_jobs(
    rows: list[tuple[int, dict[str, Any]]],
    errors: list[DesignErrorRow],
) -> dict[str, DesignJob]:
    jobs: dict[str, DesignJob] = {}
    seen_orders: dict[int, str] = {}

    for row_number, row in rows:
        job_id = _clean_text(row.get("job_id"))
        if not job_id:
            errors.append(
                DesignErrorRow(
                    sheet="jobs",
                    row_number=row_number,
                    column="job_id",
                    message="job_id is required",
                )
            )
            continue
        if job_id in jobs:
            errors.append(
                DesignErrorRow(
                    sheet="jobs",
                    row_number=row_number,
                    column="job_id",
                    message=f"duplicate job_id '{job_id}'",
                )
            )
            continue

        job_type = _clean_text(row.get("job_type"))
        if job_type not in {"merge", "migration"}:
            errors.append(
                DesignErrorRow(
                    sheet="jobs",
                    row_number=row_number,
                    column="job_type",
                    message=f"unsupported job_type '{job_type}'",
                )
            )
            job_type = job_type or ""

        order_value = _parse_positive_int(row.get("order"))
        if order_value is None:
            errors.append(
                DesignErrorRow(
                    sheet="jobs",
                    row_number=row_number,
                    column="order",
                    message="order must be a positive integer",
                )
            )
            order_value = 0
        elif order_value in seen_orders:
            errors.append(
                DesignErrorRow(
                    sheet="jobs",
                    row_number=row_number,
                    column="order",
                    message=f"duplicate order {order_value}",
                )
            )
        else:
            seen_orders[order_value] = job_id

        enabled_value, enabled_error = _parse_boolean_like(row.get("enabled"))
        if enabled_error is not None:
            errors.append(
                DesignErrorRow(
                    sheet="jobs",
                    row_number=row_number,
                    column="enabled",
                    message=enabled_error,
                )
            )
            enabled_value = False

        depends_on = _parse_depends_on(row.get("depends_on"))

        config_name = _clean_text(row.get("config_name")) or ""
        jobs[job_id] = DesignJob(
            job_id=job_id,
            job_type=job_type,
            order=order_value,
            enabled=enabled_value,
            config_name=config_name,
            depends_on=depends_on,
            row_number=row_number,
        )

    return jobs


def _validate_job_dependencies(
    jobs: dict[str, DesignJob],
    errors: list[DesignErrorRow],
) -> None:
    for job in jobs.values():
        for depends_on in job.depends_on:
            if depends_on not in jobs:
                errors.append(
                    DesignErrorRow(
                        sheet="jobs",
                        row_number=job.row_number,
                        column="depends_on",
                        message=f"depends_on references unknown job '{depends_on}'",
                    )
                )

    graph = {job_id: [dep for dep in job.depends_on if dep in jobs] for job_id, job in jobs.items()}
    visiting: set[str] = set()
    visited: set[str] = set()
    cycle: list[str] = []

    def visit(job_id: str) -> bool:
        if job_id in visited:
            return False
        if job_id in visiting:
            cycle.append(job_id)
            return True
        visiting.add(job_id)
        for depends_on in graph.get(job_id, []):
            if visit(depends_on):
                cycle.append(job_id)
                return True
        visiting.remove(job_id)
        visited.add(job_id)
        return False

    for job_id in jobs:
        if visit(job_id):
            cycle_path = list(reversed(cycle))
            if cycle_path and cycle_path[0] == cycle_path[-1]:
                cycle_text = " -> ".join(cycle_path)
            else:
                cycle_text = " -> ".join(cycle_path + [cycle_path[0]]) if cycle_path else job_id
            first_job = jobs[cycle_path[0]] if cycle_path else jobs[job_id]
            errors.append(
                DesignErrorRow(
                    sheet="jobs",
                    row_number=first_job.row_number,
                    column="depends_on",
                    message=f"dependency cycle detected: {cycle_text}",
                )
            )
            break


def _validate_detail_rows(
    sheet_rows: dict[str, list[tuple[int, dict[str, Any]]]],
    jobs: dict[str, DesignJob],
    errors: list[DesignErrorRow],
) -> None:
    detail_sheets = [sheet for sheet in REQUIRED_SHEETS if sheet not in {"project", "jobs"}]
    for sheet_name in detail_sheets:
        rows = sheet_rows.get(sheet_name, [])
        for row_number, row in rows:
            job_id = _clean_text(row.get("job_id"))
            if not job_id:
                errors.append(
                    DesignErrorRow(
                        sheet=sheet_name,
                        row_number=row_number,
                        column="job_id",
                        message="job_id is required",
                    )
                )
                continue
            if job_id not in jobs:
                errors.append(
                    DesignErrorRow(
                        sheet=sheet_name,
                        row_number=row_number,
                        column="job_id",
                        message=f"unknown job_id '{job_id}'",
                    )
                )


def _validate_required_detail_rows(
    sheet_rows: dict[str, list[tuple[int, dict[str, Any]]]],
    jobs: dict[str, DesignJob],
    errors: list[DesignErrorRow],
) -> None:
    job_rows_by_sheet: dict[str, dict[str, list[tuple[int, dict[str, Any]]]]] = {}
    for sheet_name in REQUIRED_SHEETS:
        if sheet_name in {"project", "jobs"}:
            continue
        if sheet_name not in sheet_rows:
            continue
        rows_by_job: dict[str, list[tuple[int, dict[str, Any]]]] = {}
        for row_number, row in sheet_rows.get(sheet_name, []):
            job_id = _clean_text(row.get("job_id"))
            if not job_id:
                continue
            rows_by_job.setdefault(job_id, []).append((row_number, row))
        job_rows_by_sheet[sheet_name] = rows_by_job

    required_sheet_map = {
        "merge": MERGE_REQUIRED_DETAIL_SHEETS,
        "migration": MIGRATION_REQUIRED_DETAIL_SHEETS,
    }
    for job in jobs.values():
        if not job.enabled:
            continue
        for sheet_name in required_sheet_map.get(job.job_type, []):
            if job.job_id not in job_rows_by_sheet.get(sheet_name, {}):
                errors.append(
                    DesignErrorRow(
                        sheet="jobs",
                        row_number=job.row_number,
                        column="job_id",
                        message=(
                            f"enabled {job.job_type} job requires rows in {sheet_name}"
                        ),
                    )
                )


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text or None
    return str(value).strip() or None


def _parse_positive_int(value: Any) -> int | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, int):
        return value if value > 0 else None
    if isinstance(value, float):
        if value.is_integer() and value > 0:
            return int(value)
        return None
    text = _clean_text(value)
    if text is None or not text.isdigit():
        return None
    parsed = int(text)
    return parsed if parsed > 0 else None


def _parse_boolean_like(value: Any) -> tuple[bool, str | None]:
    if isinstance(value, bool):
        return value, None
    if value is None:
        return False, "enabled must be boolean-like"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if value == 1:
            return True, None
        if value == 0:
            return False, None
        return False, "enabled must be boolean-like"
    text = _clean_text(value)
    if text is None:
        return False, "enabled must be boolean-like"
    lowered = text.lower()
    if lowered in {"true", "yes", "y", "1"}:
        return True, None
    if lowered in {"false", "no", "n", "0"}:
        return False, None
    return False, "enabled must be boolean-like"


def _parse_depends_on(value: Any) -> list[str]:
    text = _clean_text(value)
    if text is None:
        return []
    return [item.strip() for item in text.split(",") if item.strip()]
