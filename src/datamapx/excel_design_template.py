"""Build the standard DataMapX Excel design template workbook."""

from __future__ import annotations

import argparse
import csv
from collections.abc import Iterable
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SHEET_ORDER: list[str] = [
    "project",
    "jobs",
    "merge_inputs",
    "merge_rules",
    "migration_inputs",
    "input_schema",
    "references",
    "reference_schema",
    "derived",
    "outputs",
    "mappings",
    "validations",
    "filters",
    "checks",
    "error_handling",
    "runtime",
]

DEFAULT_TEMPLATE_NAME = "datamapx_design_template.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGNMENT = Alignment(vertical="center", wrap_text=True)


def _read_sheet_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open("r", encoding="utf-8", newline="") as file:
        rows = list(csv.reader(file))
    if not rows:
        raise ValueError(f"{path.name}: sheet CSV is empty")

    headers = rows[0]
    data_rows: list[list[str]] = []
    for index, row in enumerate(rows[1:], start=2):
        if not row:
            continue
        if len(row) > len(headers):
            raise ValueError(f"{path.name}:{index}: row has more columns than header")
        padded_row = row + [""] * (len(headers) - len(row))
        data_rows.append(padded_row)
    return headers, data_rows


def _apply_header_formatting(worksheet, headers: list[str], rows: list[list[str]]) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 20

    for col_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT

        max_length = len(str(header))
        for row in rows:
            value = row[col_index - 1] if col_index - 1 < len(row) else ""
            max_length = max(max_length, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_index)].width = min(max_length + 2, 40)


def build_workbook(sheets_dir: Path) -> Workbook:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_name in SHEET_ORDER:
        sheet_path = sheets_dir / f"{sheet_name}.csv"
        headers, rows = _read_sheet_csv(sheet_path)

        worksheet = workbook.create_sheet(title=sheet_name)
        worksheet.append(headers)
        for row in rows:
            worksheet.append(row)
        _apply_header_formatting(worksheet, headers, rows)

    workbook.active = 0
    return workbook


def write_template(output_path: Path, sheets_dir: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook(sheets_dir)
    workbook.save(output_path)
    return output_path


def _default_paths() -> tuple[Path, Path]:
    module_dir = Path(__file__).resolve().parents[2]
    example_dir = module_dir / "examples" / "08_excel_design"
    return example_dir / "sheets", example_dir / DEFAULT_TEMPLATE_NAME


def main(argv: Iterable[str] | None = None) -> int:
    default_sheets_dir, default_output_path = _default_paths()

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--sheets-dir",
        type=Path,
        default=default_sheets_dir,
        help="Directory that contains the sheet CSV files.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=default_output_path,
        help="Path to the generated .xlsx template.",
    )
    args = parser.parse_args(list(argv) if argv is not None else None)

    write_template(args.output, args.sheets_dir)
    print(args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
