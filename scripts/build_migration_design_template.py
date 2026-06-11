from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = Path("docs/templates/data_migration_design_template_v2.xlsx")

NAVY = "17365D"
BLUE = "2F75B5"
TEAL = "0F6B78"
LIGHT_BLUE = "D9EAF7"
LIGHT_TEAL = "DDEBF0"
LIGHT_GRAY = "E7E6E6"
PALE_GRAY = "F4F6F8"
PALE_YELLOW = "FFF2CC"
PALE_GREEN = "E2F0D9"
PALE_RED = "FCE4D6"
WHITE = "FFFFFF"
TEXT = "243447"
GRID = "A6B1BC"

THIN = Side(style="thin", color=GRID)
MEDIUM = Side(style="medium", color=NAVY)

TITLE_FONT = Font(name="Yu Gothic", size=20, bold=True, color=WHITE)
SUBTITLE_FONT = Font(name="Yu Gothic", size=11, color=WHITE)
SECTION_FONT = Font(name="Yu Gothic", size=11, bold=True, color=WHITE)
HEADER_FONT = Font(name="Yu Gothic", size=9, bold=True, color=WHITE)
BODY_FONT = Font(name="Yu Gothic", size=9, color=TEXT)
NOTE_FONT = Font(name="Yu Gothic", size=8, italic=True, color="5B6573")


def set_print_layout(
    ws,
    print_area: str,
    repeat_rows: str | None = None,
    fit_width: int = 1,
) -> None:
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = fit_width
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = print_area
    if repeat_rows:
        ws.print_title_rows = repeat_rows
    ws.sheet_properties.pageSetUpPr.autoPageBreaks = False
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.45
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    ws.oddHeader.center.text = "&B移行設計書"
    ws.oddHeader.right.text = "&A"
    ws.oddFooter.left.text = "機密区分: 社外秘"
    ws.oddFooter.center.text = "Page &P / &N"
    ws.oddFooter.right.text = "版: &F"


def title_band(ws, title: str, subtitle: str, end_col: int) -> None:
    end = get_column_letter(end_col)
    ws.merge_cells(f"A1:{end}2")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(f"A3:{end}3")
    ws["A3"] = subtitle
    ws["A3"].font = SUBTITLE_FONT
    ws["A3"].fill = PatternFill("solid", fgColor=BLUE)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 20


def section_bar(ws, row: int, title: str, end_col: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row, 1, title)
    cell.font = SECTION_FONT
    cell.fill = PatternFill("solid", fgColor=TEAL)
    cell.alignment = Alignment(vertical="center")
    cell.border = Border(top=MEDIUM, bottom=MEDIUM)
    ws.row_dimensions[row].height = 21


def style_label(cell) -> None:
    cell.font = Font(name="Yu Gothic", size=9, bold=True, color=TEXT)
    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_input(cell, center: bool = False) -> None:
    cell.font = BODY_FONT
    cell.fill = PatternFill("solid", fgColor=PALE_YELLOW)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=True,
    )
    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    cell.protection = Protection(locked=False)


def style_header_row(ws, row: int, start_col: int, headers: list[str]) -> None:
    for offset, header in enumerate(headers):
        cell = ws.cell(row, start_col + offset, header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    ws.row_dimensions[row].height = 42


def style_entry_grid(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in ws.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=end_col,
    ):
        for cell in row:
            style_input(cell)
    for row_no in range(start_row, end_row + 1):
        if row_no % 2 == 0:
            for col_no in range(start_col, end_col + 1):
                ws.cell(row_no, col_no).fill = PatternFill("solid", fgColor="FFF9E6")


def add_list_validation(ws, cell_range: str, formula: str) -> None:
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "一覧から値を選択してください。"
    validation.errorTitle = "入力値エラー"
    validation.prompt = "一覧から選択できます。"
    validation.promptTitle = "入力候補"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    ws.add_data_validation(validation)
    validation.add(cell_range)


def add_required_formatting(ws, cell_range: str, first_cell: str) -> None:
    red_fill = PatternFill("solid", fgColor=PALE_RED)
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=[f'AND({first_cell}="",COUNTA($A1:$AZ1)>0)'], fill=red_fill),
    )


def set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def create_cover(wb: Workbook) -> None:
    ws = wb.create_sheet("01_表紙")
    set_print_layout(ws, "A1:L31")
    ws.sheet_view.zoomScale = 85
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 12

    ws.merge_cells("A1:L5")
    ws["A1"] = "データ移行設計書"
    ws["A1"].font = Font(name="Yu Gothic", size=28, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for row in range(1, 6):
        ws.row_dimensions[row].height = 24

    ws.merge_cells("A6:L6")
    ws["A6"] = "Data Migration Design Specification"
    ws["A6"].font = Font(name="Yu Gothic", size=13, italic=True, color=WHITE)
    ws["A6"].fill = PatternFill("solid", fgColor=BLUE)
    ws["A6"].alignment = Alignment(horizontal="center", vertical="center")

    rows = [
        (9, "プロジェクト名", "（プロジェクト名称）"),
        (11, "対象システム", "（移行元システム → 移行先システム）"),
        (13, "移行単位／対象領域", "（業務領域・データ領域）"),
        (15, "文書番号", "MIG-DES-001"),
        (17, "版数", "0.1"),
        (19, "文書ステータス", "作成中"),
        (21, "作成日", ""),
        (23, "作成者／所属", ""),
        (25, "承認者／承認日", ""),
        (27, "機密区分", "社外秘"),
    ]
    for row, label, value in rows:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=11)
        style_label(ws.cell(row, 2))
        ws.cell(row, 2, label)
        style_input(ws.cell(row, 5))
        ws.cell(row, 5, value)
        ws.row_dimensions[row].height = 25

    add_list_validation(ws, "E19", "'99_選択肢'!$A$2:$A$5")
    add_list_validation(ws, "E27", "'99_選択肢'!$B$2:$B$5")
    ws["E21"].number_format = "yyyy/mm/dd"
    ws["E25"].number_format = "yyyy/mm/dd"
    ws.merge_cells("B30:K31")
    ws["B30"] = (
        "本書は、移行対象データ、ファイルレイアウト、変換・抽出・照合・例外処理を定義し、"
        "関係者間の合意および移行結果の検証根拠とする。"
    )
    ws["B30"].font = NOTE_FONT
    ws["B30"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def create_revision_history(wb: Workbook) -> None:
    ws = wb.create_sheet("02_改訂履歴")
    title_band(ws, "02 改訂履歴", "文書の変更内容と承認状況を記録します。", 12)
    set_print_layout(ws, "A1:L33", "6:6")
    section_bar(ws, 5, "改訂履歴", 12)
    headers = [
        "No.",
        "版数",
        "改訂日",
        "改訂区分",
        "改訂概要",
        "対象シート／章",
        "作成者",
        "レビュー者",
        "承認者",
        "承認日",
        "チケット／根拠",
        "備考",
    ]
    style_header_row(ws, 6, 1, headers)
    style_entry_grid(ws, 7, 32, 1, 12)
    for row in range(7, 33):
        ws.cell(row, 1, row - 6)
        ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 3).number_format = "yyyy/mm/dd"
        ws.cell(row, 10).number_format = "yyyy/mm/dd"
        ws.row_dimensions[row].height = 29
    add_list_validation(ws, "D7:D32", "'99_選択肢'!$C$2:$C$5")
    set_widths(
        ws,
        {
            "A": 5,
            "B": 8,
            "C": 12,
            "D": 11,
            "E": 28,
            "F": 22,
            "G": 12,
            "H": 12,
            "I": 12,
            "J": 12,
            "K": 18,
            "L": 20,
        },
    )
    ws.freeze_panes = "A7"
    ws.auto_filter.ref = "A6:L32"


def create_overview(wb: Workbook) -> None:
    ws = wb.create_sheet("03_移行概要")
    title_band(
        ws,
        "03 移行概要",
        "移行の目的、対象範囲、実施方針、品質基準および責任分担を定義します。",
        14,
    )
    set_print_layout(ws, "A1:N58")
    ws.sheet_view.zoomScale = 75
    set_widths(
        ws,
        {
            "A": 3,
            "B": 17,
            "C": 15,
            "D": 15,
            "E": 15,
            "F": 15,
            "G": 15,
            "H": 15,
            "I": 15,
            "J": 15,
            "K": 15,
            "L": 15,
            "M": 15,
            "N": 3,
        },
    )
    sections = [
        (5, "1. 目的・背景", 6, 9, "本移行の目的、背景、達成すべき業務上の状態を記載する。"),
        (
            11,
            "2. 対象範囲／対象外",
            12,
            16,
            "対象業務・データ・期間・組織、および対象外を明確に記載する。",
        ),
        (
            18,
            "3. 移行元・移行先と前提条件",
            19,
            23,
            "システム、環境、基準日時、文字コード、タイムゾーン等の前提を記載する。",
        ),
        (
            25,
            "4. 実施方式・スケジュール",
            26,
            30,
            "移行回数、リハーサル、本番切替、停止時間、再実行単位を記載する。",
        ),
        (
            32,
            "5. 品質基準・照合方針",
            33,
            37,
            "件数・金額・キー・サンプル照合、許容差、合否判定、証跡を記載する。",
        ),
        (
            39,
            "6. エラー／例外／ロールバック方針",
            40,
            44,
            "エラー分類、除外・補正・再処理、承認、復旧手順を記載する。",
        ),
        (
            46,
            "7. セキュリティ・データ管理",
            47,
            50,
            "個人情報、暗号化、授受方法、保管期限、削除、アクセス権を記載する。",
        ),
        (
            52,
            "8. 体制・責任分担・承認",
            53,
            57,
            "責任者、作業者、レビュー者、業務確認者、最終承認者を記載する。",
        ),
    ]
    for bar_row, title, start, end, prompt in sections:
        section_bar(ws, bar_row, title, 14)
        ws.merge_cells(start_row=start, start_column=2, end_row=end, end_column=13)
        cell = ws.cell(start, 2, prompt)
        style_input(cell)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.comment = Comment("記載内容を具体的な事実・判断基準に置き換えてください。", "Template")
        for row in range(start, end + 1):
            ws.row_dimensions[row].height = 21


FILE_HEADERS = [
    "ファイルID",
    "ファイル名／論理名",
    "用途・データ内容",
    "提供元／管理元",
    "授受方式・保管場所",
    "ファイル形式",
    "文字コード",
    "区切り文字",
    "ヘッダー",
    "想定件数",
    "抽出基準日時／対象期間",
    "主キー／一意性",
    "更新方式",
    "機密区分",
    "備考",
]

FIELD_HEADERS = [
    "ファイルID",
    "項目No.",
    "項目ID",
    "項目名（論理）",
    "物理項目名",
    "項目説明／業務定義",
    "データ型",
    "桁数",
    "小数桁",
    "必須",
    "キー区分",
    "書式／フォーマット",
    "許容値・コード体系",
    "NULL／空白の意味",
    "既定値",
    "編集・正規化ルール",
    "機密／個人情報区分",
    "サンプル値",
    "備考",
]


def create_format_sheet(
    wb: Workbook,
    name: str,
    title: str,
    subtitle: str,
    role_note: str,
    reference: bool = False,
    output: bool = False,
) -> None:
    ws = wb.create_sheet(name)
    title_band(ws, title, subtitle, 19)
    set_print_layout(ws, "A1:S72", "18:18", fit_width=2)
    ws.sheet_view.zoomScale = 65

    section_bar(ws, 5, "記載方針", 19)
    ws.merge_cells("A6:S7")
    ws["A6"] = role_note
    ws["A6"].font = NOTE_FONT
    ws["A6"].fill = PatternFill("solid", fgColor=PALE_GRAY)
    ws["A6"].alignment = Alignment(vertical="center", wrap_text=True)
    ws["A6"].border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    section_bar(ws, 9, "ファイル定義（複数ファイルを1行ずつ登録）", 19)
    style_header_row(ws, 10, 1, FILE_HEADERS)
    style_entry_grid(ws, 11, 15, 1, 15)
    for row in range(11, 16):
        ws.row_dimensions[row].height = 34
    add_list_validation(ws, "F11:F15", "'99_選択肢'!$D$2:$D$7")
    add_list_validation(ws, "G11:G15", "'99_選択肢'!$E$2:$E$6")
    add_list_validation(ws, "I11:I15", "'99_選択肢'!$F$2:$F$4")
    add_list_validation(ws, "N11:N15", "'99_選択肢'!$B$2:$B$5")

    section_bar(ws, 17, "項目定義（ファイルIDで対象ファイルに紐付け）", 19)
    style_header_row(ws, 18, 1, FIELD_HEADERS)
    style_entry_grid(ws, 19, 69, 1, 19)
    for row in range(19, 70):
        ws.row_dimensions[row].height = 31
    add_list_validation(ws, "G19:G69", "'99_選択肢'!$G$2:$G$12")
    add_list_validation(ws, "J19:J69", "'99_選択肢'!$F$2:$F$4")
    add_list_validation(ws, "K19:K69", "'99_選択肢'!$H$2:$H$6")
    add_list_validation(ws, "Q19:Q69", "'99_選択肢'!$I$2:$I$7")

    if reference:
        ws["A6"] = (
            "移行判断や値変換に使用する参照データを定義する。参照時点、キーの一意性、"
            "重複時の採用ルール、有効期間、未一致時の扱いを必ず明示する。"
        )
        for cell_ref in ("L11", "C11"):
            ws[cell_ref].comment = Comment(
                "参照キーの一意性と、重複・未一致時の扱いを備考欄にも記載してください。",
                "Template",
            )
    elif output:
        ws["A6"] = (
            "移行先へ登録・連携する成果物のレイアウトを定義する。移行先の正式な項目定義、"
            "必須制約、桁・型・コード体系、出力順、受入条件と一致させる。"
        )
    else:
        ws["A6"] = (
            "移行元から受領するファイルの実体を定義する。抽出条件や変換内容ではなく、"
            "受領時点のレイアウト、項目の意味、品質特性、NULL・空白の意味を記載する。"
        )

    set_widths(
        ws,
        {
            "A": 12,
            "B": 9,
            "C": 12,
            "D": 18,
            "E": 23,
            "F": 30,
            "G": 13,
            "H": 9,
            "I": 9,
            "J": 9,
            "K": 12,
            "L": 18,
            "M": 26,
            "N": 24,
            "O": 14,
            "P": 28,
            "Q": 18,
            "R": 18,
            "S": 24,
        },
    )
    ws.freeze_panes = "F19"
    ws.auto_filter.ref = "A18:S69"
    ws.merge_cells("A71:S72")
    ws["A71"] = (
        "確認ポイント: ファイルIDの重複、項目No.の順序、物理項目名の重複、必須・キー・型・桁・"
        "コード体系・NULL定義の未記載がないこと。"
    )
    ws["A71"].font = NOTE_FONT
    ws["A71"].fill = PatternFill("solid", fgColor=PALE_GREEN)
    ws["A71"].alignment = Alignment(vertical="center", wrap_text=True)


def create_transform_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("07_データ変換方式")
    title_band(
        ws,
        "07 データ変換方式",
        "抽出・結合・変換・例外処理・検証を、移行元から移行先まで追跡可能な形で定義します。",
        24,
    )
    set_print_layout(ws, "A1:X112", "19:19", fit_width=3)
    ws.sheet_view.zoomScale = 55

    section_bar(ws, 5, "1. 共通変換・処理方針", 24)
    common_headers = [
        "方針ID",
        "適用範囲",
        "処理区分",
        "処理内容／判定基準",
        "実施順序",
        "エラー時処理",
        "検証方法",
        "承認者",
    ]
    style_header_row(ws, 6, 1, common_headers)
    style_entry_grid(ws, 7, 13, 1, 8)
    for row in range(7, 14):
        ws.row_dimensions[row].height = 34
    add_list_validation(ws, "C7:C13", "'99_選択肢'!$J$2:$J$13")
    add_list_validation(ws, "F7:F13", "'99_選択肢'!$K$2:$K$7")

    section_bar(ws, 15, "2. 項目マッピング・変換仕様", 24)
    mapping_headers = [
        "ルールID",
        "処理順",
        "対象出力ファイルID",
        "出力項目ID",
        "出力物理項目名",
        "移行区分",
        "入力ファイルID",
        "入力項目ID／物理名",
        "参照ファイルID",
        "参照キー／条件",
        "取得項目",
        "抽出条件",
        "変換区分",
        "変換式・判定条件・処理内容",
        "コード変換／対応表",
        "型・桁・書式変換",
        "NULL／空白処理",
        "既定値／固定値",
        "複数一致時の優先順位",
        "未一致・異常時処理",
        "出力例",
        "検証方法／期待結果",
        "根拠資料／要件ID",
        "備考",
    ]
    style_header_row(ws, 19, 1, mapping_headers)
    style_entry_grid(ws, 20, 89, 1, 24)
    for row in range(20, 90):
        ws.row_dimensions[row].height = 42
    add_list_validation(ws, "F20:F89", "'99_選択肢'!$L$2:$L$8")
    add_list_validation(ws, "M20:M89", "'99_選択肢'!$J$2:$J$13")
    add_list_validation(ws, "T20:T89", "'99_選択肢'!$K$2:$K$7")

    ws["A16"] = "前提"
    style_label(ws["A16"])
    ws.merge_cells("B16:X17")
    ws["B16"] = (
        "1出力項目につき1行を基本とし、条件分岐が複数ある場合はルールIDを分ける。"
        "『そのまま』『適宜変換』のような解釈依存の記載は禁止し、入力値・条件・出力値を明示する。"
    )
    style_input(ws["B16"])
    ws["B16"].alignment = Alignment(vertical="center", wrap_text=True)

    section_bar(ws, 91, "3. コード変換・値対応表", 24)
    code_headers = [
        "対応表ID",
        "対象ルールID",
        "入力ファイルID",
        "入力項目",
        "入力値",
        "入力値名称",
        "出力値",
        "出力値名称",
        "有効開始日",
        "有効終了日",
        "未定義値の扱い",
        "根拠／備考",
    ]
    style_header_row(ws, 92, 1, code_headers)
    style_entry_grid(ws, 93, 102, 1, 12)
    for row in range(93, 103):
        ws.row_dimensions[row].height = 31

    section_bar(ws, 104, "4. 照合・受入判定", 24)
    recon_headers = [
        "検証ID",
        "対象",
        "検証観点",
        "移行前集計／期待値",
        "移行後集計／実績値",
        "比較方法",
        "許容差",
        "合格条件",
        "証跡",
        "確認者",
        "結果",
        "備考",
    ]
    style_header_row(ws, 105, 1, recon_headers)
    style_entry_grid(ws, 106, 111, 1, 12)
    add_list_validation(ws, "K106:K111", "'99_選択肢'!$M$2:$M$5")

    widths = {
        "A": 12,
        "B": 8,
        "C": 15,
        "D": 14,
        "E": 23,
        "F": 13,
        "G": 14,
        "H": 24,
        "I": 14,
        "J": 28,
        "K": 18,
        "L": 28,
        "M": 15,
        "N": 42,
        "O": 28,
        "P": 25,
        "Q": 24,
        "R": 18,
        "S": 22,
        "T": 24,
        "U": 20,
        "V": 28,
        "W": 20,
        "X": 24,
    }
    set_widths(ws, widths)
    ws.freeze_panes = "F20"
    ws.auto_filter.ref = "A19:X89"


def create_lists(wb: Workbook) -> None:
    ws = wb.create_sheet("99_選択肢")
    lists = {
        "A": ["文書ステータス", "作成中", "レビュー中", "承認済", "廃止"],
        "B": ["機密区分", "公開", "社内限定", "社外秘", "極秘"],
        "C": ["改訂区分", "新規", "変更", "訂正", "廃止"],
        "D": ["ファイル形式", "CSV", "TSV", "固定長", "Excel", "XML", "その他"],
        "E": ["文字コード", "UTF-8", "UTF-8 BOM", "Shift_JIS", "CP932", "その他"],
        "F": ["有無", "有", "無", "対象外"],
        "G": [
            "データ型",
            "文字列",
            "整数",
            "小数",
            "日付",
            "日時",
            "時刻",
            "真偽値",
            "バイナリ",
            "コード",
            "その他",
        ],
        "H": ["キー区分", "主キー", "複合主キー", "外部キー", "一意キー", "非キー"],
        "I": ["情報区分", "一般", "社内限定", "個人情報", "要配慮個人情報", "機密", "匿名化済"],
        "J": [
            "処理区分",
            "直接転記",
            "固定値",
            "条件分岐",
            "文字列編集",
            "数値計算",
            "日付変換",
            "コード変換",
            "参照取得",
            "集約",
            "分割",
            "除外",
            "その他",
        ],
        "K": [
            "異常時処理",
            "エラー停止",
            "対象行除外",
            "警告継続",
            "既定値設定",
            "NULL設定",
            "個別判断",
        ],
        "L": [
            "移行区分",
            "移行",
            "非移行",
            "新規生成",
            "参照のみ",
            "削除対象",
            "要個別判断",
            "対象外",
        ],
        "M": ["検証結果", "未実施", "合格", "不合格", "条件付合格"],
    }
    for col, values in lists.items():
        for row, value in enumerate(values, start=1):
            ws[f"{col}{row}"] = value
    ws.sheet_state = "veryHidden"


def create_workbook() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    create_lists(wb)
    create_cover(wb)
    create_revision_history(wb)
    create_overview(wb)
    create_format_sheet(
        wb,
        "04_入力フォーマット",
        "04 入力用フォーマット",
        "移行元から受領する入力ファイルと、その項目仕様を定義します。",
        "",
    )
    create_format_sheet(
        wb,
        "05_参照フォーマット",
        "05 参照用フォーマット",
        "変換・判定に利用する参照ファイルと、そのキー・値仕様を定義します。",
        "",
        reference=True,
    )
    create_format_sheet(
        wb,
        "06_出力フォーマット",
        "06 出力用フォーマット",
        "移行先へ受け渡す出力ファイルと、その受入レイアウトを定義します。",
        "",
        output=True,
    )
    create_transform_sheet(wb)
    wb._sheets.append(wb._sheets.pop(0))

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.security.lockStructure = False
    wb.properties.title = "データ移行設計書テンプレート"
    wb.properties.subject = "データ移行のレビュー・承認・実施・検証に使用する標準テンプレート"
    wb.properties.creator = "Data Migration Office"
    wb.properties.description = (
        "特定の移行ツールに依存せず、移行設計そのものを定義するためのExcelテンプレート。"
    )

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)


def verify_workbook() -> None:
    wb = load_workbook(OUTPUT, data_only=False)
    expected = [
        "01_表紙",
        "02_改訂履歴",
        "03_移行概要",
        "04_入力フォーマット",
        "05_参照フォーマット",
        "06_出力フォーマット",
        "07_データ変換方式",
        "99_選択肢",
    ]
    if wb.sheetnames != expected:
        raise RuntimeError(f"Unexpected sheet order: {wb.sheetnames}")
    for ws in wb.worksheets[:-1]:
        if ws.page_setup.orientation != "landscape":
            raise RuntimeError(f"{ws.title}: orientation is not landscape")
        if str(ws.page_setup.paperSize) != str(ws.PAPERSIZE_A4):
            raise RuntimeError(f"{ws.title}: paper size is not A4")
        if not ws.print_area:
            raise RuntimeError(f"{ws.title}: print area is missing")
    if wb["99_選択肢"].sheet_state != "veryHidden":
        raise RuntimeError("Selection-list sheet must be hidden")


if __name__ == "__main__":
    create_workbook()
    verify_workbook()
    print(OUTPUT.resolve())
